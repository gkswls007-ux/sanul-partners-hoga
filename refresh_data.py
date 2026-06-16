import json
from pathlib import Path
from datetime import date, datetime

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\gkswl\OneDrive\바탕 화면\호가 및 실거래가_V2.xlsx")
OUTPUT = Path(__file__).parent / "data" / "listings.json"
SHEET_NAME = "세대별정리_row"
BROKER_SHEET_NAME = "동일매물중개사"

KEYMAP = {
    "조사일자": "surveyDate",
    "아파트단지명": "complex",
    "구분": "dealType",
    "공급면적": "supplyArea",
    "전용면적": "exclusiveArea",
    "평수(공급)": "pyeong",
    "동": "building",
    "층수": "floor",
    "층구분": "floorGroup",
    "매매호가/보증금": "price",
    "월세": "monthlyRent",
    "환산 보증금": "convertedDeposit",
    "매물 특징": "features",
    "중개사 수": "brokerCount",
    "평수묶음": "pyeongGroup",
    "평당매매가": "pricePerPyeong",
    "입주가능일": "moveIn",
    "방향": "direction",
    "대표매물번호": "representativeListingId",
}


def clean_value(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return value


def clean_listing_id(value):
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_broker_map(workbook):
    if BROKER_SHEET_NAME not in workbook.sheetnames:
        return {}

    sheet = workbook[BROKER_SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(header).strip() for header in next(rows)]
    broker_map = {}

    for row in rows:
        item = dict(zip(headers, row))
        representative_id = clean_listing_id(item.get("대표매물번호"))
        individual_id = clean_listing_id(item.get("개별매물번호"))
        broker_name = str(item.get("공인중개사사무소명") or "").strip()
        if not representative_id or not broker_name:
            continue

        broker_map.setdefault(representative_id, [])
        broker = next(
            (item for item in broker_map[representative_id] if item["brokerName"] == broker_name),
            None,
        )
        if broker is None:
            broker = {
                "brokerName": broker_name,
                "individualListingIds": [],
            }
            broker_map[representative_id].append(broker)
        if individual_id and individual_id not in broker["individualListingIds"]:
            broker["individualListingIds"].append(individual_id)

    return broker_map


def main():
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(header).strip() for header in next(rows)]
    data = []
    broker_map = load_broker_map(workbook)

    for row in rows:
        item = {}
        has_value = False
        for header, value in zip(headers, row):
            if value not in (None, ""):
                has_value = True
            key = KEYMAP.get(header, header)
            if key == "representativeListingId":
                item[key] = clean_listing_id(value)
            else:
                item[key] = clean_value(value)
        if has_value:
            data.append(item)

    OUTPUT.parent.mkdir(exist_ok=True)
    OUTPUT.write_text(
        json.dumps(
            {
                "meta": {
                    "source": f"{SOURCE.name} / {SHEET_NAME}",
                    "rowCount": len(data),
                    "headers": headers,
                    "brokerMatchCount": len(broker_map),
                },
                "rows": data,
                "brokerMap": broker_map,
            },
            ensure_ascii=False,
            separators=(",", ":"),
        ),
        encoding="utf-8",
    )
    print(f"{len(data):,}건 업데이트 완료: {OUTPUT}")


if __name__ == "__main__":
    main()
