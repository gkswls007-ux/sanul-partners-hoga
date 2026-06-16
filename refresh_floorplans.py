import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).parent
MANIFEST = ROOT / "data" / "floorplans.json"

SOURCES = [
    {
        "source": Path(r"C:\Users\gkswl\OneDrive\바탕 화면\2단지_자이더시티도면.xlsx"),
        "complex": "산울2단지세종자이더시티",
        "folder": "sanul2",
    },
    {
        "source": Path(r"C:\Users\gkswl\OneDrive\바탕 화면\5단지_산울파밀리에더파크.xlsx"),
        "complex": "산울5단지세종파밀리에더파크",
        "folder": "sanul5",
    },
    {
        "source": Path(r"C:\Users\gkswl\OneDrive\바탕 화면\도면_산울6단지세종리첸시아파밀리에(주상복함).xlsx"),
        "complex": "산울6단지세종리첸시아파밀리에(주상복합)",
        "folder": "sanul6",
    },
    {
        "source": Path(r"C:\Users\gkswl\OneDrive\바탕 화면\산울7단지세종리첸시아파밀리에.xlsx"),
        "complex": "산울7단지세종리첸시아파밀리에(주상복합)",
        "folder": "sanul7",
    },
    {
        "source": Path(r"C:\Users\gkswl\OneDrive\바탕 화면\해밀1단지마스터힐스.xlsx"),
        "complex": "해밀1단지마스터힐스",
        "folder": "haemil1",
    },
    {
        "source": Path(r"C:\Users\gkswl\OneDrive\바탕 화면\해밀마을2단지.xlsx"),
        "complex": "해밀2단지마스터힐스",
        "folder": "haemil2",
    },
    {
        "source": Path(r"C:\Users\gkswl\OneDrive\바탕 화면\산울8단지엘리프세종.xlsx"),
        "complex": "산울8단지엘리프세종6-3",
        "folder": "sanul8",
    },
]


def clean_filename(value):
    return re.sub(r"[^0-9A-Za-z가-힣_-]+", "_", str(value)).strip("_")


def collect_texts(sheet):
    texts = []
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if value not in (None, ""):
                texts.append(str(value).strip())
    return texts


def value_after_label(texts, label):
    for index, value in enumerate(texts):
        if value == label and index + 1 < len(texts):
            return texts[index + 1]
    return ""


def image_sort_key(image):
    anchor = getattr(image, "anchor", None)
    try:
        return (anchor._from.row, anchor._from.col)
    except Exception:
        return (0, 0)


def collect_labeled_cells(sheet):
    labels = []
    for row in sheet.iter_rows():
        for cell in row:
            value = str(cell.value).strip() if cell.value not in (None, "") else ""
            if value in {"1층", "2층", "1F", "2F"}:
                labels.append({"label": value, "row": cell.row, "col": cell.column})
    return labels


def label_for_image(image, labels, fallback):
    try:
        image_row = image.anchor._from.row + 1
        image_col = image.anchor._from.col + 1
    except Exception:
        return fallback

    if not labels:
        return fallback

    same_side = [
        item
        for item in labels
        if item["row"] <= image_row and abs(item["col"] - image_col) <= 3
    ]
    candidates = same_side or [item for item in labels if item["row"] <= image_row] or labels
    best = min(candidates, key=lambda item: (abs(item["col"] - image_col), abs(item["row"] - image_row)))
    return best["label"]


def label_order(label):
    return {"1층": 1, "1F": 1, "2층": 2, "2F": 2}.get(label, 99)


def main():
    MANIFEST.parent.mkdir(exist_ok=True)

    existing_plans = {}
    if MANIFEST.exists():
        existing_plans = json.loads(MANIFEST.read_text(encoding="utf-8")).get("plans", {})

    plans = {}
    for config in SOURCES:
        if not config["source"].exists():
            plans.update(
                {
                    key: value
                    for key, value in existing_plans.items()
                    if value.get("complex") == config["complex"]
                }
            )
            print(f"원본 파일 없음, 기존 도면 유지: {config['source']}")
            continue

        workbook = load_workbook(config["source"], data_only=True)
        output_dir = ROOT / "assets" / "floorplans" / config["folder"]
        output_dir.mkdir(parents=True, exist_ok=True)

        for old_file in output_dir.glob("*"):
            if old_file.is_file():
                old_file.unlink()

        for sheet in workbook.worksheets:
            type_name = sheet.title
            texts = collect_texts(sheet)
            labels = collect_labeled_cells(sheet)
            household_count = value_after_label(texts, "해당면적세대수")
            rooms_baths = value_after_label(texts, "방수/욕실수")
            images = []
            image_labels = []

            image_items = []
            for image in sorted(sheet._images, key=image_sort_key):
                label = label_for_image(image, labels, f"{len(image_items) + 1}F")
                image_items.append((image, label))
            if len(image_items) > 1:
                image_items.sort(key=lambda item: (label_order(item[1]), image_sort_key(item[0])))

            for index, (image, label) in enumerate(image_items, start=1):
                extension = image.path.split(".")[-1].lower() if getattr(image, "path", "") else "png"
                if extension not in {"png", "jpg", "jpeg"}:
                    extension = "png"
                suffix = f"_{index}" if len(sheet._images) > 1 else ""
                filename = f"{clean_filename(type_name)}{suffix}.{extension}"
                target = output_dir / filename
                target.write_bytes(image._data())
                images.append(f"./assets/floorplans/{config['folder']}/{filename}")
                image_labels.append(label)

            plans[f"{config['complex']}|{type_name}"] = {
                "complex": config["complex"],
                "type": type_name,
                "structure": "복층형" if len(images) > 1 else "일반형",
                "householdCount": household_count,
                "roomsBaths": rooms_baths,
                "images": images,
                "imageLabels": image_labels,
            }

    MANIFEST.write_text(json.dumps({"plans": plans}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"도면 {len(plans):,}개 타입 추출 완료: {MANIFEST}")


if __name__ == "__main__":
    main()
