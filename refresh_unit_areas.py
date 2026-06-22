import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\gkswl\OneDrive\바탕 화면\동호수별면적.xlsx")
OUTPUT = Path(__file__).parent / "data" / "unit_areas.json"
SHEET_NAME = "동호수_평형정보"

KEYMAP = {
    "조사일자": "surveyDate",
    "아파트단지명": "complex",
    "단지번호": "complexNumber",
    "동": "building",
    "동번호": "buildingNumber",
    "호수": "unit",
    "층": "floor",
    "라인": "line",
    "타입번호": "typeNumber",
    "공급면적": "supplyArea",
    "타입명": "typeName",
    "전용면적": "exclusiveArea",
    "평수(공급)": "pyeong",
    "필로티여부": "piloti",
}


def clean_value(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    return value


def clean_text(value):
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def main():
    if not SOURCE.exists():
        print(f"원본 파일 없음, 기존 동호수별 면적 유지: {SOURCE}")
        return

    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        print(f"시트 없음, 기존 동호수별 면적 유지: {SHEET_NAME}")
        return

    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(header).strip() for header in next(rows)]
    data = []

    for row in rows:
        item = {}
        has_value = False
        for header, value in zip(headers, row):
            if value not in (None, ""):
                has_value = True
            key = KEYMAP.get(header, header)
            item[key] = clean_value(value)
        if not has_value:
            continue
        item["complex"] = clean_text(item.get("complex"))
        item["building"] = clean_text(item.get("building"))
        item["unit"] = clean_text(item.get("unit"))
        item["typeName"] = clean_text(item.get("typeName"))
        item["supplyArea"] = clean_text(item.get("supplyArea"))
        if item["complex"] and item["building"] and item["unit"]:
            data.append(item)

    OUTPUT.parent.mkdir(exist_ok=True)
    OUTPUT.write_text(
        json.dumps(
            {
                "meta": {
                    "source": f"{SOURCE.name} / {SHEET_NAME}",
                    "rowCount": len(data),
                    "headers": headers,
                },
                "rows": data,
            },
            ensure_ascii=False,
            separators=(",", ":"),
        ),
        encoding="utf-8",
    )
    print(f"{len(data):,}건 동호수별 면적 업데이트 완료: {OUTPUT}")


if __name__ == "__main__":
    main()
