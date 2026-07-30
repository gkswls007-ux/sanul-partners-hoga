import json
from pathlib import Path

from openpyxl import load_workbook

from refresh_data import KEYMAP, REAL_TRANSACTION_KEYMAP, clean_listing_id, clean_value


SOURCE = Path(r"C:\Users\gkswl\OneDrive\바탕 화면\호가 및 실거래가_수원.xlsx")
OUTPUT = Path(__file__).parent / "data" / "listings-suwon.json"
LISTING_SHEET = "매물요약"
BROKER_SHEET = "동일매물_중개사"
REAL_TRANSACTION_SHEET = "타입별_실거래가"
SUWON_KEYMAP = {**KEYMAP, "입주유형": "moveIn"}
VALID_DEAL_TYPES = {"매매", "전세", "월세"}


def is_valid_listing_row(row):
    return (
        bool(row.get("surveyDate"))
        and bool(row.get("complex"))
        and row.get("dealType") in VALID_DEAL_TYPES
    )


def read_rows(sheet, keymap, listing_ids=False):
    rows = sheet.iter_rows(values_only=True)
    headers = [str(header).strip() for header in next(rows)]
    data = []

    for row in rows:
        item = {}
        has_value = False
        for header, value in zip(headers, row):
            if value not in (None, ""):
                has_value = True
            key = keymap.get(header, header)
            item[key] = clean_listing_id(value) if listing_ids and key == "representativeListingId" else clean_value(value)
        if has_value:
            data.append(item)
    return headers, data


def load_broker_map(workbook):
    if BROKER_SHEET not in workbook.sheetnames:
        return {}

    rows = workbook[BROKER_SHEET].iter_rows(values_only=True)
    headers = [str(header).strip() for header in next(rows)]
    broker_map = {}

    for row in rows:
        item = dict(zip(headers, row))
        representative_id = clean_listing_id(item.get("대표매물번호"))
        individual_id = clean_listing_id(item.get("개별매물번호"))
        broker_name = str(item.get("공인중개사사무소명") or "").strip()
        if not representative_id or not broker_name:
            continue

        broker_map.setdefault(representative_id, [])
        broker = next((entry for entry in broker_map[representative_id] if entry["brokerName"] == broker_name), None)
        if broker is None:
            broker = {"brokerName": broker_name, "individualListingIds": []}
            broker_map[representative_id].append(broker)
        if individual_id and individual_id not in broker["individualListingIds"]:
            broker["individualListingIds"].append(individual_id)

    return broker_map


def merge_listing_history(incoming_rows):
    if not OUTPUT.exists():
        return incoming_rows

    try:
        existing_rows = json.loads(OUTPUT.read_text(encoding="utf-8")).get("rows", [])
    except (OSError, json.JSONDecodeError):
        return incoming_rows

    incoming_dates = {row.get("surveyDate") for row in incoming_rows if row.get("surveyDate")}
    preserved_rows = [
        row
        for row in existing_rows
        if is_valid_listing_row(row) and row.get("surveyDate") not in incoming_dates
    ]
    return preserved_rows + incoming_rows


def main():
    if not SOURCE.exists():
        print(f"수원 원본 파일 없음, 기존 수원 데이터 유지: {SOURCE}")
        return

    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    required_sheets = [LISTING_SHEET, BROKER_SHEET, REAL_TRANSACTION_SHEET]
    missing_sheets = [name for name in required_sheets if name not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"수원 원본에 필요한 시트가 없습니다: {', '.join(missing_sheets)}")

    headers, source_rows = read_rows(workbook[LISTING_SHEET], SUWON_KEYMAP, listing_ids=True)
    incoming_rows = [row for row in source_rows if is_valid_listing_row(row)]
    skipped_row_count = len(source_rows) - len(incoming_rows)
    rows = merge_listing_history(incoming_rows)
    _, real_transactions = read_rows(workbook[REAL_TRANSACTION_SHEET], REAL_TRANSACTION_KEYMAP)
    broker_map = load_broker_map(workbook)

    OUTPUT.parent.mkdir(exist_ok=True)
    OUTPUT.write_text(
        json.dumps(
            {
                "meta": {
                    "region": "수원",
                    "source": f"{SOURCE.name} / {LISTING_SHEET}",
                    "rowCount": len(rows),
                    "currentRowCount": len(incoming_rows),
                    "skippedRowCount": skipped_row_count,
                    "headers": headers,
                    "brokerMatchCount": len(broker_map),
                    "realTransactionCount": len(real_transactions),
                },
                "rows": rows,
                "brokerMap": broker_map,
                "realTransactions": real_transactions,
            },
            ensure_ascii=False,
            separators=(",", ":"),
        ),
        encoding="utf-8",
    )
    skipped_note = f", 비매물 {skipped_row_count:,}건 제외" if skipped_row_count else ""
    print(
        f"수원 {len(incoming_rows):,}건 반영{skipped_note}, "
        f"누적 {len(rows):,}건: {OUTPUT}"
    )


if __name__ == "__main__":
    main()
